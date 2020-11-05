#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2019/9/10 10:58
# @Author : xuancai
# @Email : 444@tontisa.com
# @File : read_excel.py
# @Software: PyCharm

import openpyxl

class Case:
    # 这个类用来存储用例,用例的数据设为对象的属性
    def __init__(self, attrs):
        """
        初始化用例
        :param attrs: zip类型-->[(key,value),(key1,value1)....]
        """
        for item in attrs:
            setattr(self, item[0], item[1])

class ReadExcel(object):
    """
    读写excel
    """
    def __init__(self,file_name,sheet_name):
        """
        初始化用例数据
        :param file_name:  文件名称  -->  str
        :param sheet_name: 表单名称  -->  str
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """
        打开工作簿,选择表单
        :return:
        """
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet =self.wb[self.sheet_name]

    def close(self):
        """
        关闭工作簿
        :return:
        """
        self.wb.close()

    def read_data_obj(self):
        self.open()
        """ 
        按行读取excel中的数据，以列表的形式返回，列表中每个对象为一条用例
        excel中的表头为对象的属性，对应的数据为属性值
        zip(打包组合)
        :return: type:list--->[case_obj1,case_obj2....]，
        """
        #按行获取数据转换成列表
        rows_data =list(self.sheet.rows)
        #获取表单表头数据
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        #判断是否读取所有数据
        for case in rows_data[1:]:
            #data用例临时存放用例数据
            data = []
            #判断该单元格是否为字符串
            for cell in case:
                data.append(cell.value)
            #将该条数据加入cases中
            case_data = zip(titles,data)
            #创建一个Case类的对象，用来保存用列数据
            case_obj = Case(case_data)
            cases.append(case_obj)
            print(case_obj.data)
        self.close()
        return cases


    def readline_data_obj(self,list1=None):
        """
        按指定的列，读取excel中数据，以列表的形式返回，列表每一个对象为一条用例
        excel中的表头为对象属性，对应的数据为属性的值
        :param list1: list  -->要读取的列[1,2....]
        :return: type:list--->[case_obj1,case_obj2....]，
        """
        self.open()
        list1 = eval(list1)
        if len(list1) == 0:
            return self.read_data_obj()
        max_r = self.sheet.max_row
        #定义一个空列表，用来存放用例
        cases = []
        #定义一个空列表来存放表头
        titles = []
        #遍历出所有的行
        for row in range(1,max_r+1):
            #判断是否是第一行
            if row == 1:
                #遍历list1指定的列，获取表头
                for column in list1:
                    title = self.sheet.cell(row,column).value
                    #将数据加入表头
                    titles.append(title)
            else:
                #定义一个空列表来存放该行的数据
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row,column).value
                    case_data.append(info)
                #将该条数据与表头进行打包组合
                case = list(zip(titles,case_data))
                # 创建一个用例对象，将表头和数据传进入初始化
                case_obj = Case(case)
                cases.append(case_obj)
            self.close()
            return cases

    def write_data(self,row,column,msg):
        self.open()
        #写入数据
        self.sheet.cell(row=row,column=column,value=msg)
        self.wb.save(self.file_name)
        self.close()




